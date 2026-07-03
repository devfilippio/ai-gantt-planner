from __future__ import annotations

import io
import re

from openpyxl import Workbook, load_workbook

from api.models import Plan, Task
from api.scheduler import CycleError, compute_schedule

HEADERS = ["задача", "описание", "исполнитель", "длительность", "предшественники"]


class ImportError_(ValueError):
    pass


def _pad_row(row: tuple, width: int = 5) -> tuple:
    """A reviewer's spreadsheet may have fewer columns than the spec (or trailing
    columns trimmed by openpyxl) — pad with None instead of crashing on row[3]."""
    if len(row) >= width:
        return row
    return tuple(row) + (None,) * (width - len(row))


def _slug(name: str, taken: set[str]) -> str:
    base = re.sub(r"[^a-z0-9]+", "-", name.lower()).strip("-") or "task"
    slug, i = base, 1
    while slug in taken:
        i += 1
        slug = f"{base}-{i}"
    taken.add(slug)
    return slug


def export_plan(plan: Plan) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "План"
    ws.append(HEADERS)
    name_by_id = {t.id: t.name for t in plan.tasks}
    for t in plan.tasks:
        preds = ", ".join(name_by_id.get(p, p) for p in t.predecessors)
        ws.append([t.name, t.description, t.assignee, t.duration_days, preds])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def import_plan(data: bytes) -> Plan:
    try:
        wb = load_workbook(io.BytesIO(data))
    except Exception:
        raise ImportError_("Не удалось прочитать файл — ожидается корректный .xlsx")
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ImportError_("Пустой файл")
    taken: set[str] = set()
    raw = []
    name_to_id: dict[str, str] = {}
    for idx, row in enumerate(rows[1:], start=2):
        row = _pad_row(row)
        name = (row[0] or "").strip() if row[0] else ""
        if not name:
            continue
        try:
            duration = int(row[3])
        except (TypeError, ValueError):
            raise ImportError_(f"строка {idx}: некорректная длительность '{row[3]}'")
        if duration <= 0:
            raise ImportError_(f"строка {idx}: длительность должна быть > 0")
        tid = _slug(name, taken)
        name_to_id[name] = tid
        raw.append((idx, tid, name, row))
    tasks = []
    for idx, tid, name, row in raw:
        pred_names = [p.strip() for p in str(row[4] or "").replace(";", ",").split(",") if p.strip()]
        preds = []
        for pn in pred_names:
            if pn not in name_to_id:
                raise ImportError_(f"строка {idx}: предшественник '{pn}' не найден")
            if name_to_id[pn] == tid:
                raise ImportError_(f"строка {idx}: задача '{name}' не может зависеть от себя")
            preds.append(name_to_id[pn])
        tasks.append(Task(
            id=tid, name=name, description=str(row[1] or ""),
            assignee=str(row[2] or ""), duration_days=int(row[3]), predecessors=preds,
        ))
    plan = Plan(tasks=tasks)
    # Reject dependency cycles HERE, before the caller persists the plan —
    # otherwise a cyclic upload gets saved and every subsequent schedule
    # computation (GET /api/plan included) crashes: the whole app looks dead
    # until a manual reset.
    try:
        compute_schedule(plan)
    except CycleError as e:
        names = {t.id: t.name for t in plan.tasks}
        cycle = " → ".join(names.get(i, i) for i in e.cycle)
        raise ImportError_(f"В файле цикл зависимостей: {cycle}")
    return plan
