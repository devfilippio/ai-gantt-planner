import io

import pytest

from api.models import Task, Plan
from api.excel import export_plan, import_plan, HEADERS
from api.excel import ImportError_ as ExcelImportError


def _plan():
    return Plan(tasks=[
        Task(id="a", name="Design", description="mockups", assignee="Maria", duration_days=3, predecessors=[]),
        Task(id="b", name="API", description="auth", assignee="Ivan", duration_days=5, predecessors=["a"]),
    ], project_start="2026-05-05")


def test_export_produces_readable_workbook():
    from openpyxl import load_workbook
    data = export_plan(_plan())
    wb = load_workbook(io.BytesIO(data))
    ws = wb.active
    headers = [c.value for c in ws[1]]
    assert headers == ["задача", "описание", "исполнитель", "длительность", "предшественники"]
    assert ws[2][0].value == "Design"
    assert ws[3][4].value == "Design"  # predecessor referenced by NAME


def test_import_roundtrips_export():
    original = _plan()
    data = export_plan(original)
    imported = import_plan(data)
    assert [t.name for t in imported.tasks] == ["Design", "API"]
    # predecessor name resolved back to an id
    api_task = next(t for t in imported.tasks if t.name == "API")
    design_task = next(t for t in imported.tasks if t.name == "Design")
    assert api_task.predecessors == [design_task.id]


def test_import_unknown_predecessor_reports_row():
    from openpyxl import Workbook
    wb = Workbook(); ws = wb.active
    ws.append(HEADERS)
    ws.append(["A", "", "X", 2, ""])
    ws.append(["B", "", "Y", 3, "Ghost"])
    buf = io.BytesIO(); wb.save(buf)
    with pytest.raises(ExcelImportError) as exc:
        import_plan(buf.getvalue())
    assert "3" in str(exc.value)  # row number
    assert "Ghost" in str(exc.value)


def test_import_rejects_dependency_cycle_before_save():
    """A reviewer uploading a cyclic plan must get a clean 400-style error —
    previously the cyclic plan was saved and every subsequent schedule
    computation (the whole app) crashed with 500 until a manual reset."""
    from openpyxl import Workbook
    import io as _io
    wb = Workbook(); ws = wb.active
    ws.append(HEADERS)
    ws.append(["A", "", "X", 2, "B"])
    ws.append(["B", "", "Y", 3, "A"])
    buf = _io.BytesIO(); wb.save(buf)
    with pytest.raises(ExcelImportError) as exc:
        import_plan(buf.getvalue())
    assert "цикл" in str(exc.value).lower()


def test_import_rejects_self_reference():
    from openpyxl import Workbook
    import io as _io
    wb = Workbook(); ws = wb.active
    ws.append(HEADERS)
    ws.append(["A", "", "X", 2, "A"])
    buf = _io.BytesIO(); wb.save(buf)
    with pytest.raises(ExcelImportError) as exc:
        import_plan(buf.getvalue())
    assert "себя" in str(exc.value)


def test_import_tolerates_missing_columns():
    """Sheets with fewer than 5 columns must produce a row-level error, not an
    IndexError/500."""
    from openpyxl import Workbook
    import io as _io
    wb = Workbook(); ws = wb.active
    ws.append(["задача", "описание", "исполнитель"])
    ws.append(["A", "", ""])
    buf = _io.BytesIO(); wb.save(buf)
    with pytest.raises(ExcelImportError) as exc:
        import_plan(buf.getvalue())
    assert "строка 2" in str(exc.value)


def test_import_garbage_zip_gives_clean_error():
    with pytest.raises(ExcelImportError):
        import_plan(b"PK\x03\x04 this is not a real xlsx")
